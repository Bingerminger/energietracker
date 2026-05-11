#!/usr/bin/env python3
"""
init_data.py — Excel-Importer für Energietracker v1.0.0

Liest data/input.xlsx mit den Tabs "Gas" und "Strom" und schreibt
Zähler und Readings in das neue v1.0.0-Datenmodell. Cumulative Import:
bereits vorhandene Readings werden nicht dupliziert.

Erwartetes Excel-Format pro Tab (Gas, Strom):
    Spalte A: Datum (DD.MM.YYYY oder Excel-Datum)
    Spalte B: Stand (Zahl)
    Spalte C: Notiz (optional)
    Spalte D: Auffälligkeit (optional, 1/0 oder "ja"/"nein")

Wasser wird über die UI angelegt — der Use-Case eines Excel-Imports
für Wasser ist zu selten, um den Code dafür zu pflegen.

Aufruf:
    pip install -r requirements.txt
    python scripts/init_data.py
"""

from __future__ import annotations

import json
import os
import sys
import uuid
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

try:
    from openpyxl import load_workbook
except ImportError:
    print("Fehler: openpyxl nicht installiert. Bitte 'pip install -r requirements.txt' ausführen.", file=sys.stderr)
    sys.exit(1)


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
INPUT_XLSX = DATA_DIR / "input.xlsx"

UTILITIES = ["gas", "strom"]


# ---------------------------------------------------------------------------
# Hilfsfunktionen
# ---------------------------------------------------------------------------

def new_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:10]}"


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    tmp.replace(path)


def parse_date(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def parse_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def parse_bool(value: Any) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"1", "ja", "yes", "true", "x"}
    return False


# ---------------------------------------------------------------------------
# Hauptlogik
# ---------------------------------------------------------------------------

def ensure_meter(utility: str) -> dict:
    """Gibt den Default-Zähler für die Utility zurück, legt ihn an falls nötig."""
    meters_path = DATA_DIR / utility / "meters.json"
    meters = read_json(meters_path, [])

    if meters:
        return meters[0]

    meter = {
        "id": new_id("m"),
        "name": "Hauptzähler",
        "icon": "🔥" if utility == "gas" else "⚡",
        "created_at": datetime.now().date().isoformat(),
        "active": True,
        "notes": "",
        "devices": [
            {
                "id": new_id("d"),
                "serial": None,
                "installed_on": "2000-01-01",
                "initial_counter": 0.0,
                "removed_on": None,
                "final_counter": None,
                "reason": None,
            }
        ],
    }
    meters.append(meter)
    write_json(meters_path, meters)
    return meter


def device_id_for_date(meter: dict, iso_date: str) -> str:
    """Liefert das Device, das am gegebenen Datum aktiv war."""
    target = datetime.fromisoformat(iso_date).date()
    for device in meter["devices"]:
        installed = datetime.fromisoformat(device["installed_on"]).date() if device.get("installed_on") else None
        removed = datetime.fromisoformat(device["removed_on"]).date() if device.get("removed_on") else None
        if installed and installed > target:
            continue
        if removed and removed < target:
            continue
        return device["id"]
    # Fallback: erstes Device
    return meter["devices"][0]["id"]


def import_sheet(utility: str, sheet) -> tuple[int, int]:
    """Importiert ein Excel-Sheet in die Utility. Liefert (added, skipped)."""
    meter = ensure_meter(utility)
    readings_path = DATA_DIR / utility / "readings.json"
    readings = read_json(readings_path, [])

    # Index existierender Readings nach Datum
    existing_dates = {r["date"] for r in readings if r.get("meter_id") == meter["id"]}

    added = 0
    skipped = 0

    rows = list(sheet.iter_rows(values_only=True))
    # erste Zeile als Header annehmen, wenn die erste Zelle nicht-numerisch ist
    if rows and rows[0] and isinstance(rows[0][0], str) and parse_date(rows[0][0]) is None:
        rows = rows[1:]

    for raw_row in rows:
        if not raw_row or raw_row[0] is None:
            continue

        iso_date = parse_date(raw_row[0])
        value = parse_float(raw_row[1]) if len(raw_row) > 1 else None
        note = raw_row[2] if len(raw_row) > 2 else ""
        anomaly = parse_bool(raw_row[3]) if len(raw_row) > 3 else False

        if iso_date is None or value is None:
            skipped += 1
            continue

        if iso_date in existing_dates:
            skipped += 1
            continue

        readings.append({
            "id": new_id("r"),
            "meter_id": meter["id"],
            "device_id": device_id_for_date(meter, iso_date),
            "date": iso_date,
            "counter": value,
            "price_cents": None,
            "note": (note or "").strip() if isinstance(note, str) else "",
            "is_estimated": anomaly,
            "is_future": iso_date > datetime.now().date().isoformat(),
        })
        existing_dates.add(iso_date)
        added += 1

    # Sortieren nach Datum aufsteigend
    readings.sort(key=lambda r: r["date"])
    write_json(readings_path, readings)

    return added, skipped


def main() -> int:
    if not INPUT_XLSX.exists():
        print(f"Fehler: {INPUT_XLSX} nicht gefunden.", file=sys.stderr)
        print("Bitte input.xlsx im data/-Verzeichnis platzieren.", file=sys.stderr)
        return 1

    print(f"Lese {INPUT_XLSX} ...")
    wb = load_workbook(INPUT_XLSX, data_only=True)

    total_added = 0
    total_skipped = 0

    for utility in UTILITIES:
        sheet_name = next((n for n in wb.sheetnames if n.lower() == utility), None)
        if sheet_name is None:
            print(f"  [{utility}] Kein Tab gefunden — übersprungen.")
            continue

        sheet = wb[sheet_name]
        added, skipped = import_sheet(utility, sheet)
        total_added += added
        total_skipped += skipped
        print(f"  [{utility}] {added} neu hinzugefügt, {skipped} übersprungen (bereits vorhanden oder unvollständig).")

    # meta.json aktualisieren falls noch nicht vorhanden
    meta_path = DATA_DIR / "meta.json"
    meta = read_json(meta_path, {})
    if not meta:
        meta = {
            "schema_version": "1.0.0",
            "migrated_at": datetime.now().isoformat(timespec="seconds"),
            "log": [{"step": "init_data.py", "added": total_added}],
        }
        write_json(meta_path, meta)

    print(f"\nFertig. {total_added} neue Readings, {total_skipped} übersprungen.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
